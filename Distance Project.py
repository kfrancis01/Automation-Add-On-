#!/usr/bin/env python3
# Record, plot, and execute commands depending on distance inputs 

import serial
import RPi.GPIO as GPIO
import sys
import time
import matplotlib.pyplot as plt
import datetime as dt
import matplotlib.animation as animation
import numpy as np
import openpyxl

#Imports for creating xlxs data file
from openpyxl import Workbook
from openpyxl.styles import Color, Side, PatternFill, Font, Border, Alignment

# Rename and grab the active worksheet
wb = Workbook()
ws = wb.active

# Create and format .xlxs file
ws['A1'] = dt.datetime.now()
ws.column_dimensions['A'].width = 20
ws.merge_cells('B2:D2')

ws['B2'].font = Font(italic = True, bold = True,size = 20)
ws['B2'] = 'Recorded Distances'
ws['B2'].border = Border(bottom = Side(border_style ='thin'))
ws['B3'].font = Font(italic = True, bold = True,size = 14)
ws['C3'].font = Font(italic = True, bold = True,size = 14)
ws['D3'].font = Font(italic = True, bold = True,size = 14)
ws.merge_cells('B3:D3')
ws['B3'] = 'Trial One'
ws['B3'].border = Border(bottom = Side(border_style ='thin'))
ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
ws.column_dimensions['B'].width = len('Trial One')

i = 1

#Set USB serial communication to Arduino
ard = serial.Serial('/dev/ttyACM0',57600, timeout = 1)

#create Figure
fig = plt.figure(1)

#Create subplots for figure one
ax1 = fig.add_subplot(3, 1, 1)
ax2 = fig.add_subplot(3, 1, 3)

#Create empty matrices for distance and increment
dist = []
inc = []
            
#Create function to animate plots
def animate(x):
    global i
    global dist
    global inc
    ard = serial.Serial('/dev/ttyACM0',57600, timeout = 1)
    input = ard.readline()
    time.sleep(0.5)
    if input != b'':
        read = int(input)
    
    #Add to distance and increment matrices
        dist.append(read)
        inc.append(i)
    #Create each axes
        ax1.clear()
        ax2.clear()
        plt.subplot(3,1,1)
        plt.ylabel('Distance (Inches)')
        plt.subplot(3,1,3)
        plt.ylabel('Distance (Inches)')
        plt.figure(1)
        ax1.set_title('Distance Over Time')
        ax1.set_xlabel('Increments')
   
   #Create plot
        ax1.plot(inc,dist, color = 'blue', linestyle = 'dashed', marker = 'o')
        ax1.grid(True)
    
        fig.suptitle('Distance of Puck', fontsize = 16) 

   #Create bar graph
        ax2.bar(0,read,width = 0.1)
        ax2.grid(True)
        ax2.xaxis.set_ticks(np.arange(0))
        ax2.yaxis.set_ticks((0,10,20,30,40,50,60,70,80,90,100))
        ax2.set_ylim(0,100)

        print(input)
        
#if read is not equal to 1000 then the trip wire has been activated
        if read != 1000:
            ws.append(['',i,read,'inches'])
            i = i +1
        elif read == 2:
            print("Zone Two") #place desired command here
        elif read == 3:
            print("Zone Three")
        elif read == 4:
            print("Zone Four")
        elif read == 5:
            print("Zone Five")
        elif read == 6:
            print("Zone Six")

        wb.save("DistanceRecord.xlsx")

#start the animation process
ani = animation.FuncAnimation(fig, animate, interval=200)
plt.draw()
plt.show()

GPIO.cleanup()

